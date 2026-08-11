import io
import os
import tempfile
import unittest
import zipfile
from datetime import datetime, timezone

from openpyxl import load_workbook

from nexus_trade.exports import ExportSafetyError, ReportExporter
from nexus_trade.reports import ReportService
from nexus_trade.scheduler import BrasiliaSchedule
from tests.test_nexus_trade_reports import report_evidence


class ReportExportTests(unittest.TestCase):
    def make_report(self):
        self.temp = tempfile.TemporaryDirectory()
        service = ReportService(os.path.join(self.temp.name, "reports.db"))
        evidence = report_evidence()
        evidence["daily"][0]["note"] = "=HYPERLINK(\"https://evil.test\")"
        evidence["audit"][0]["detail"] = "+cmd|' /C calc'!A0"
        window = BrasiliaSchedule().weekly_window(
            datetime(2026, 8, 10, 13, tzinfo=timezone.utc)
        )
        return service.close_weekly(window, evidence)

    def tearDown(self):
        if hasattr(self, "temp"):
            self.temp.cleanup()

    def test_csv_zip_has_exact_tables_and_neutralizes_formula_cells_deterministically(self):
        report = self.make_report()
        exporter = ReportExporter()

        first = exporter.csv_zip(report)
        second = exporter.csv_zip(report)

        self.assertEqual(first, second)
        with zipfile.ZipFile(io.BytesIO(first)) as archive:
            self.assertEqual(
                set(archive.namelist()),
                {"summary.csv", "days.csv", "champion.csv", "trial.csv", "indicators.csv", "gates.csv", "audit.csv"},
            )
            self.assertIn("'=HYPERLINK", archive.read("days.csv").decode("utf-8-sig"))
            self.assertIn("'+cmd", archive.read("audit.csv").decode("utf-8-sig"))
            self.assertTrue(all(".." not in name and not name.startswith(("/", "\\")) for name in archive.namelist()))

    def test_xlsx_has_exact_sheets_stable_values_no_formulas_or_macros(self):
        report = self.make_report()
        payload = ReportExporter().xlsx(report)

        workbook = load_workbook(io.BytesIO(payload), data_only=False, keep_vba=False)
        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "Days", "Champion", "Trial", "Indicators", "Gates", "Audit"],
        )
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    self.assertNotEqual(cell.data_type, "f")
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            self.assertFalse(any("vba" in name.lower() for name in archive.namelist()))

    def test_export_rejects_secret_fields_absolute_paths_and_unsafe_envelopes(self):
        exporter = ReportExporter()
        for unsafe in (
            {"api_token": "secret"},
            {"audit": [{"detail": "C:\\Users\\operator\\report.json"}]},
            {"window": {"path": "../escape"}},
        ):
            with self.subTest(unsafe=unsafe):
                with self.assertRaises(ExportSafetyError):
                    exporter.csv_zip(unsafe)


if __name__ == "__main__":
    unittest.main()
